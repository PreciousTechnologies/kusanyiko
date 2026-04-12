from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from django.db.models import Count, Q
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from datetime import timedelta, datetime
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from members.models import Member
from users.models import User, AuditLog
from .models import ExportHistory, BrandingSettings
from .serializers import BrandingSettingsSerializer


KANDA_AREAS = {
    'dar_es_salaam_na_pwani': [
        'Dar es Salaam', 'Pwani', 'Mwenge', 'Temeke', 'Ushindi', 'Imara',
        'Kinondoni', 'Zanzibar', 'Yombo', 'Kisukulu', 'Kisukuru'
    ],
    'nyanda_za_juu_kusini': ['Mbeya', 'Rukwa', 'Katavi', 'Iringa'],
    'kusini': ['Mtwara', 'Lindi', 'Ruvuma', 'Njombe'],
    'kaskazini': ['Kilimanjaro', 'Arusha', 'Manyara', 'Tanga'],
    'magharibi_na_ziwa': ['Kigoma', 'Shinyanga', 'Simiyu', 'Mwanza', 'Geita', 'Mara', 'Kagera', 'Tabora'],
    'kati': ['Dodoma', 'Singida'],
}


def get_member_scope_queryset(user):
    queryset = Member.objects.filter(is_deleted=False)

    if user.role == 'admin':
        return queryset

    if user.role == 'apostle':
        if not user.kanda:
            return queryset.none()

        areas = KANDA_AREAS.get(user.kanda, [])
        scope_q = Q()
        for area in areas:
            scope_q |= Q(region__iexact=area)
            scope_q |= Q(center_area__iexact=area)
            scope_q |= Q(zone__iexact=area)

        return queryset.filter(scope_q)

    return queryset.filter(created_by=user)


def _safe_percent(part, total):
    if not total:
        return 0.0
    return round((part / total) * 100, 1)


def _members_for_period(date_range):
    queryset = Member.objects.filter(is_deleted=False)

    start_date = parse_date(str(date_range.get('start_date', '')).strip()) if date_range else None
    end_date = parse_date(str(date_range.get('end_date', '')).strip()) if date_range else None

    if start_date:
        queryset = queryset.filter(created_at__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__date__lte=end_date)

    return queryset


def _create_workbook(title):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    return workbook, sheet


def _style_title(sheet, text):
    sheet['A1'] = text
    sheet.merge_cells('A1:D1')
    sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
    sheet['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    sheet.row_dimensions[1].height = 28


def _style_table_header(row):
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin = Side(style='thin', color='C9CED6')
    for cell in row:
        cell.font = Font(bold=True, color='1F2937')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _autosize_columns(sheet, min_width=14, max_width=46):
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        longest = max((len(v) for v in values), default=min_width)
        width = max(min_width, min(longest + 2, max_width))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _build_kpi_section(sheet, row_idx, members_qs):
    total_members = members_qs.count()
    male_count = members_qs.filter(gender='male').count()
    female_count = members_qs.filter(gender='female').count()
    saved_count = members_qs.filter(saved=True).count()
    unsaved_count = members_qs.filter(saved=False).count()

    sheet.cell(row=row_idx, column=1, value='Executive KPI Summary').font = Font(size=12, bold=True, color='1F4E78')
    row_idx += 1

    headers = ['Indicator', 'Value', 'Share (%)']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=row_idx, column=col, value=header)
    _style_table_header(sheet[row_idx])
    row_idx += 1

    rows = [
        ('Total Members Registered', total_members, 100.0 if total_members else 0.0),
        ('Male Members', male_count, _safe_percent(male_count, total_members)),
        ('Female Members', female_count, _safe_percent(female_count, total_members)),
        ('Saved Members', saved_count, _safe_percent(saved_count, total_members)),
        ('Unsaved Members', unsaved_count, _safe_percent(unsaved_count, total_members)),
    ]
    for label, value, percentage in rows:
        sheet.cell(row=row_idx, column=1, value=label)
        sheet.cell(row=row_idx, column=2, value=value)
        sheet.cell(row=row_idx, column=3, value=percentage)
        sheet.cell(row=row_idx, column=3).number_format = '0.0'
        row_idx += 1

    return row_idx + 1


def _write_distribution_sheet(workbook, title, header_a, rows):
    ws = workbook.create_sheet(title=title)
    _style_title(ws, title)
    ws['A3'] = header_a
    ws['B3'] = 'Member Count'
    _style_table_header(ws[3])

    current_row = 4
    for row_label, count in rows:
        ws.cell(row=current_row, column=1, value=row_label or 'Not Specified')
        ws.cell(row=current_row, column=2, value=count)
        current_row += 1

    ws.freeze_panes = 'A4'
    _autosize_columns(ws)


def _workbook_response(workbook, filename):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name='ExecutiveTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.white,
            leading=22,
            spaceAfter=0,
        )
    )
    styles.add(
        ParagraphStyle(
            name='ExecutiveMeta',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#374151'),
            leading=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name='ExecutiveSection',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#1F4E78'),
            leading=14,
            spaceBefore=10,
            spaceAfter=6,
        )
    )
    return styles


def _pdf_header_block(report_title, subtitle):
    title_table = Table([[Paragraph(report_title, _pdf_styles()['ExecutiveTitle'])]], colWidths=[18.5 * cm])
    title_table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1F4E78')),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]
        )
    )

    meta_style = _pdf_styles()['ExecutiveMeta']
    metadata = [
        Paragraph(f'<b>Prepared On:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', meta_style),
        Paragraph(f'<b>Report Scope:</b> {subtitle}', meta_style),
    ]

    return [title_table, Spacer(1, 0.25 * cm), *metadata, Spacer(1, 0.25 * cm)]


def _pdf_table(rows, col_widths=None, repeat_rows=1):
    table = Table(rows, colWidths=col_widths, repeatRows=repeat_rows)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#C9CED6')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _pdf_response(story, filename, pagesize=A4):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response


class BrandingSettingsView(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get(self, request):
        # Singleton record with pk=1 keeps deployment and querying simple.
        branding, _ = BrandingSettings.objects.get_or_create(pk=1)
        return Response(BrandingSettingsSerializer(branding).data)

    def patch(self, request):
        if not request.user.is_authenticated or getattr(request.user, 'role', '') != 'admin':
            return Response({'error': 'Unauthorized'}, status=403)

        branding, _ = BrandingSettings.objects.get_or_create(pk=1)
        serializer = BrandingSettingsSerializer(branding, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class AdminStatsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        if request.user.role != 'admin':
            return Response({'error': 'Unauthorized'}, status=403)
        
        # Get basic stats
        total_members = Member.objects.filter(is_deleted=False).count()
        
        # Get breakdown by country
        country_stats = Member.objects.filter(is_deleted=False)\
            .values('country')\
            .annotate(count=Count('id'))\
            .order_by('-count')
        
        # Get breakdown by region
        region_stats = Member.objects.filter(is_deleted=False)\
            .values('region')\
            .annotate(count=Count('id'))\
            .order_by('-count')
        
        # Get breakdown by gender
        gender_stats = Member.objects.filter(is_deleted=False)\
            .values('gender')\
            .annotate(count=Count('id'))
        
        # Get breakdown by marital status
        marital_stats = Member.objects.filter(is_deleted=False)\
            .values('marital_status')\
            .annotate(count=Count('id'))
        
        # Get saved vs unsaved
        saved_stats = Member.objects.filter(is_deleted=False)\
            .values('saved')\
            .annotate(count=Count('id'))
        
        # Get recent registrations (last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_registrations = Member.objects.filter(
            is_deleted=False,
            created_at__gte=thirty_days_ago
        ).count()
        
        # Get weekly growth data (last 8 weeks)
        weekly_data = []
        for i in range(8):
            week_start = timezone.now() - timedelta(weeks=i+1)
            week_end = timezone.now() - timedelta(weeks=i)
            week_count = Member.objects.filter(
                is_deleted=False,
                created_at__gte=week_start,
                created_at__lt=week_end
            ).count()
            weekly_data.insert(0, {
                'week': f'Week {8-i}',
                'count': week_count
            })
        
        return Response({
            'total_members': total_members,
            'country_stats': list(country_stats),
            'region_stats': list(region_stats),
            'gender_stats': list(gender_stats),
            'marital_stats': list(marital_stats),
            'saved_stats': list(saved_stats),
            'recent_registrations': recent_registrations,
            'weekly_growth': weekly_data,
        })


class RegistrantStatsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        scoped_members = get_member_scope_queryset(request.user)

        # Get basic stats for current user
        total_registered = scoped_members.count()
        
        # Get breakdown by gender
        gender_stats = scoped_members.values('gender').annotate(count=Count('id'))
        
        # Get breakdown by region
        region_stats = scoped_members.values('region').annotate(count=Count('id'))
        
        # Get saved vs unsaved
        saved_stats = scoped_members.values('saved').annotate(count=Count('id'))
        
        # Get recent registrations (last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_registrations = scoped_members.filter(created_at__gte=thirty_days_ago).count()
        
        # Get weekly performance (last 4 weeks)
        weekly_data = []
        for i in range(4):
            week_start = timezone.now() - timedelta(weeks=i+1)
            week_end = timezone.now() - timedelta(weeks=i)
            week_count = scoped_members.filter(created_at__gte=week_start, created_at__lt=week_end).count()
            weekly_data.insert(0, {
                'week': f'Week {4-i}',
                'count': week_count
            })
        
        # Get recent activity (last 5 members)
        recent_members = scoped_members.order_by('-created_at')[:5].values(
            'first_name', 'last_name', 'created_at'
        )
        
        return Response({
            'total_registered': total_registered,
            'gender_stats': list(gender_stats),
            'region_stats': list(region_stats),
            'saved_stats': list(saved_stats),
            'recent_registrations': recent_registrations,
            'weekly_performance': weekly_data,
            'recent_activity': list(recent_members),
        })


# Export Views

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_members(request):
    """Export members data in various formats"""
    format_type = request.data.get('format', 'csv')
    filters = request.data.get('filters', {})
    
    # Build queryset with filters
    queryset = get_member_scope_queryset(request.user)
    
    if filters.get('created_by') and request.user.role == 'admin':
        queryset = queryset.filter(created_by=filters['created_by'])
    
    if filters.get('region'):
        queryset = queryset.filter(region=filters['region'])
    
    if filters.get('gender'):
        queryset = queryset.filter(gender=filters['gender'])
    
    if filters.get('date_from'):
        queryset = queryset.filter(created_at__gte=filters['date_from'])
    
    if filters.get('date_to'):
        queryset = queryset.filter(created_at__lte=filters['date_to'])
    
    # Export based on format
    if format_type == 'csv':
        response = export_members_csv(queryset)
    elif format_type == 'excel':
        response = export_members_excel(queryset)
    elif format_type == 'pdf':
        response = export_members_pdf(queryset)
    else:
        return Response({'error': 'Unsupported format'}, status=400)
    
    # Log export activity
    file_size = f"{len(response.content) / 1024:.1f} KB"
    ExportHistory.objects.create(
        export_type='members',
        format=format_type,
        created_by=request.user,
        file_size=file_size,
        filters_applied=filters
    )
    
    return response


def export_members_csv(queryset):
    """Export members to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="members_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'First Name', 'Last Name', 'Email', 'Phone', 'Country', 'Region', 
        'Gender', 'Marital Status', 'Saved', 'Date Registered', 'Registered By'
    ])
    
    for member in queryset:
        writer.writerow([
            member.first_name,
            member.last_name,
            member.email,
            member.mobile_no,
            member.country,
            member.region,
            member.gender,
            member.marital_status,
            'Yes' if member.saved else 'No',
            member.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            member.created_by.username if member.created_by else 'N/A'
        ])
    
    return response


def export_members_excel(queryset):
    """Export members to a professionally formatted Excel workbook."""
    workbook, sheet = _create_workbook('Member Register')
    _style_title(sheet, 'KUSANYIKO MEMBER REGISTER')

    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet['A3'] = 'Generated On'
    sheet['B3'] = generated_at
    sheet['A4'] = 'Records'
    sheet['B4'] = queryset.count()
    sheet['A3'].font = Font(bold=True)
    sheet['A4'].font = Font(bold=True)

    headers = [
        'First Name', 'Last Name', 'Email', 'Phone', 'Country', 'Region',
        'Gender', 'Marital Status', 'Saved', 'Date Registered', 'Registered By'
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=col, value=header)
    _style_table_header(sheet[6])

    row_idx = 7
    for member in queryset:
        values = [
            member.first_name,
            member.last_name,
            member.email,
            member.mobile_no,
            member.country,
            member.region,
            member.gender,
            member.marital_status,
            'Yes' if member.saved else 'No',
            member.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            member.created_by.username if member.created_by else 'N/A'
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    sheet.freeze_panes = 'A7'
    _autosize_columns(sheet)

    filename = f'members-register-{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return _workbook_response(workbook, filename)


def export_members_pdf(queryset):
    """Export members to a branded and professional PDF report."""
    styles = _pdf_styles()
    total_members = queryset.count()
    saved_count = queryset.filter(saved=True).count()
    male_count = queryset.filter(gender='male').count()
    female_count = queryset.filter(gender='female').count()

    story = []
    story.extend(_pdf_header_block('KUSANYIKO MEMBER REGISTER', f'Total records: {total_members}'))

    story.append(Paragraph('Executive KPI Summary', styles['ExecutiveSection']))
    kpi_rows = [
        ['Indicator', 'Value', 'Share (%)'],
        ['Total Members Registered', total_members, '100.0' if total_members else '0.0'],
        ['Male Members', male_count, f'{_safe_percent(male_count, total_members):.1f}'],
        ['Female Members', female_count, f'{_safe_percent(female_count, total_members):.1f}'],
        ['Saved Members', saved_count, f'{_safe_percent(saved_count, total_members):.1f}'],
    ]
    story.append(_pdf_table(kpi_rows, col_widths=[9.5 * cm, 3 * cm, 3 * cm]))
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph('Detailed Member Register', styles['ExecutiveSection']))
    member_rows = [[
        'First Name', 'Last Name', 'Email', 'Phone', 'Country', 'Region',
        'Gender', 'Marital Status', 'Saved', 'Registered', 'Registered By'
    ]]
    for member in queryset:
        member_rows.append([
            member.first_name,
            member.last_name,
            member.email,
            member.mobile_no,
            member.country,
            member.region,
            member.gender,
            member.marital_status,
            'Yes' if member.saved else 'No',
            member.created_at.strftime('%Y-%m-%d'),
            member.created_by.username if member.created_by else 'N/A',
        ])

    story.append(
        _pdf_table(
            member_rows,
            col_widths=[1.4 * cm, 1.4 * cm, 2.7 * cm, 1.8 * cm, 1.5 * cm, 1.8 * cm, 1.2 * cm, 1.8 * cm, 1.0 * cm, 1.9 * cm, 1.8 * cm],
        )
    )

    filename = f'members-register-{datetime.now().strftime("%Y-%m-%d")}.pdf'
    return _pdf_response(story, filename, pagesize=landscape(A4))


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_analytics(request):
    """Export analytics data"""
    try:
        format_type = request.data.get('format', 'pdf')
        export_type = request.data.get('type', 'overview')
        date_range = request.data.get('date_range', {})
        
        print(f"Export analytics: format={format_type}, type={export_type}, date_range={date_range}")
        
        if format_type not in ['pdf', 'excel']:
            return Response({'error': f'Unsupported format: {format_type}'}, status=400)
        
        if export_type not in ['overview', 'summary', 'demographics', 'geographical', 'monthly']:
            return Response({'error': f'Unsupported export type: {export_type}'}, status=400)
        
        if format_type == 'pdf':
            response = export_analytics_pdf(export_type, date_range)
        elif format_type == 'excel':
            response = export_analytics_excel(export_type, date_range)
        
        # Log export activity
        try:
            file_size = f"{len(response.content) / 1024:.1f} KB"
            ExportHistory.objects.create(
                export_type='analytics',
                format=format_type,
                created_by=request.user,
                file_size=file_size,
                filters_applied={'type': export_type, 'date_range': date_range}
            )
        except Exception as log_error:
            print(f"Failed to log export: {log_error}")
            # Don't fail the export if logging fails
        
        return response
        
    except Exception as e:
        print(f"Export analytics error: {e}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


def export_analytics_pdf(export_type, date_range):
    """Export analytics to a structured and branded executive PDF report."""
    members_qs = _members_for_period(date_range)
    styles = _pdf_styles()
    start_label = date_range.get('start_date') or 'Beginning'
    end_label = date_range.get('end_date') or 'Present'

    story = []
    story.extend(
        _pdf_header_block(
            f'KUSANYIKO {export_type.upper()} ANALYTICS REPORT',
            f'Reporting window: {start_label} to {end_label}',
        )
    )

    total_members = members_qs.count()
    males = members_qs.filter(gender='male').count()
    females = members_qs.filter(gender='female').count()
    saved = members_qs.filter(saved=True).count()
    unsaved = members_qs.filter(saved=False).count()

    story.append(Paragraph('Executive KPI Summary', styles['ExecutiveSection']))
    story.append(
        _pdf_table(
            [
                ['Indicator', 'Value', 'Share (%)'],
                ['Total Members Registered', total_members, '100.0' if total_members else '0.0'],
                ['Male Members', males, f'{_safe_percent(males, total_members):.1f}'],
                ['Female Members', females, f'{_safe_percent(females, total_members):.1f}'],
                ['Saved Members', saved, f'{_safe_percent(saved, total_members):.1f}'],
                ['Unsaved Members', unsaved, f'{_safe_percent(unsaved, total_members):.1f}'],
            ],
            col_widths=[9.5 * cm, 3 * cm, 3 * cm],
        )
    )
    story.append(Spacer(1, 0.3 * cm))
    
    if export_type in ['summary', 'overview']:
        story.append(Paragraph('Country Distribution', styles['ExecutiveSection']))
        country_stats = members_qs.values('country').annotate(count=Count('id')).order_by('-count')
        country_rows = [['Country', 'Members']]
        for stat in country_stats:
            country_rows.append([stat['country'] or 'Not Specified', stat['count']])
        story.append(_pdf_table(country_rows, col_widths=[10 * cm, 5.5 * cm]))
        story.append(Spacer(1, 0.2 * cm))
        
        story.append(Paragraph('Regional Distribution (Tanzania)', styles['ExecutiveSection']))
        tanzania_regions = members_qs.filter(country='Tanzania').exclude(region__isnull=True).exclude(region='').values('region').annotate(count=Count('id')).order_by('-count')
        region_rows = [['Region', 'Members']]
        for stat in tanzania_regions:
            region_rows.append([stat['region'] or 'Not Specified', stat['count']])
        story.append(_pdf_table(region_rows, col_widths=[10 * cm, 5.5 * cm]))
        story.append(Spacer(1, 0.2 * cm))
        
        story.append(Paragraph('Dar es Salaam Center/Area Distribution', styles['ExecutiveSection']))
        dar_areas = members_qs.filter(country='Tanzania', region='Dar es Salaam').exclude(center_area__isnull=True).exclude(center_area='').values('center_area').annotate(count=Count('id')).order_by('-count')
        area_rows = [['Center/Area', 'Members']]
        for stat in dar_areas:
            area_rows.append([stat['center_area'] or 'Not Specified', stat['count']])
        story.append(_pdf_table(area_rows, col_widths=[10 * cm, 5.5 * cm]))
    
    elif export_type == 'demographics':
        story.append(Paragraph('Marital Status Distribution', styles['ExecutiveSection']))
        marital_stats = members_qs.values('marital_status').annotate(count=Count('id')).order_by('-count')
        marital_rows = [['Marital Status', 'Members']]
        for stat in marital_stats:
            marital_rows.append([stat['marital_status'] or 'Not Specified', stat['count']])
        story.append(_pdf_table(marital_rows, col_widths=[10 * cm, 5.5 * cm]))
    
    elif export_type == 'geographical':
        story.append(Paragraph('Country Distribution', styles['ExecutiveSection']))
        country_stats = members_qs.values('country').annotate(count=Count('id')).order_by('-count')
        country_rows = [['Country', 'Members']]
        for stat in country_stats:
            country_rows.append([stat['country'] or 'Not Specified', stat['count']])
        story.append(_pdf_table(country_rows, col_widths=[10 * cm, 5.5 * cm]))
        story.append(Spacer(1, 0.2 * cm))
        
        story.append(Paragraph('Regional Distribution (Tanzania)', styles['ExecutiveSection']))
        tanzania_regions = members_qs.filter(country='Tanzania').exclude(region__isnull=True).exclude(region='').values('region').annotate(count=Count('id')).order_by('-count')
        region_rows = [['Region', 'Members']]
        for stat in tanzania_regions:
            region_rows.append([stat['region'] or 'Not Specified', stat['count']])
        story.append(_pdf_table(region_rows, col_widths=[10 * cm, 5.5 * cm]))
        story.append(Spacer(1, 0.2 * cm))
        
        story.append(Paragraph('Dar es Salaam Center/Area Distribution', styles['ExecutiveSection']))
        dar_areas = members_qs.filter(country='Tanzania', region='Dar es Salaam').exclude(center_area__isnull=True).exclude(center_area='').values('center_area').annotate(count=Count('id')).order_by('-count')
        area_rows = [['Center/Area', 'Members']]
        for stat in dar_areas:
            area_rows.append([stat['center_area'] or 'Not Specified', stat['count']])
        story.append(_pdf_table(area_rows, col_widths=[10 * cm, 5.5 * cm]))

    elif export_type == 'monthly':
        story.append(Paragraph('Monthly Registration Trend (Last 12 Months)', styles['ExecutiveSection']))
        monthly_rows = [['Month', 'Registrations']]

        current = timezone.now().date().replace(day=1)
        months = []
        for _ in range(12):
            months.insert(0, current)
            if current.month == 1:
                current = current.replace(year=current.year - 1, month=12)
            else:
                current = current.replace(month=current.month - 1)

        for month_start in months:
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)

            monthly_count = members_qs.filter(
                created_at__date__gte=month_start,
                created_at__date__lt=month_end,
            ).count()
            monthly_rows.append([month_start.strftime('%b %Y'), monthly_count])

        story.append(_pdf_table(monthly_rows, col_widths=[10 * cm, 5.5 * cm]))

    filename = f'{export_type}-report-{datetime.now().strftime("%Y-%m-%d")}.pdf'
    return _pdf_response(story, filename)


def export_analytics_excel(export_type, date_range):
    """Export analytics to a structured, executive-style Excel workbook."""
    members_qs = _members_for_period(date_range)
    report_title = f'{export_type.title()} Analytics Report'
    workbook, summary_sheet = _create_workbook('Executive Summary')
    _style_title(summary_sheet, report_title.upper())

    summary_sheet['A3'] = 'Prepared On'
    summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    summary_sheet['A4'] = 'Reporting Window'
    start_label = date_range.get('start_date') or 'Beginning'
    end_label = date_range.get('end_date') or 'Present'
    summary_sheet['B4'] = f'{start_label} to {end_label}'
    summary_sheet['A3'].font = Font(bold=True)
    summary_sheet['A4'].font = Font(bold=True)

    next_row = _build_kpi_section(summary_sheet, 6, members_qs)

    if export_type in ['summary', 'overview', 'demographics']:
        marital_stats = members_qs.values('marital_status').annotate(count=Count('id')).order_by('-count')
        summary_sheet.cell(row=next_row, column=1, value='Marital Status Distribution').font = Font(size=12, bold=True, color='1F4E78')
        next_row += 1
        summary_sheet.cell(row=next_row, column=1, value='Status')
        summary_sheet.cell(row=next_row, column=2, value='Member Count')
        _style_table_header(summary_sheet[next_row])
        next_row += 1
        for stat in marital_stats:
            summary_sheet.cell(row=next_row, column=1, value=stat['marital_status'] or 'Not Specified')
            summary_sheet.cell(row=next_row, column=2, value=stat['count'])
            next_row += 1

    country_stats = list(members_qs.values('country').annotate(count=Count('id')).order_by('-count'))
    tanzania_regions = list(
        members_qs.filter(country='Tanzania')
        .exclude(region__isnull=True)
        .exclude(region='')
        .values('region')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    dar_areas = list(
        members_qs.filter(country='Tanzania', region='Dar es Salaam')
        .exclude(center_area__isnull=True)
        .exclude(center_area='')
        .values('center_area')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    if export_type in ['summary', 'overview', 'geographical']:
        _write_distribution_sheet(
            workbook,
            'Countries',
            'Country',
            [(stat['country'] or 'Not Specified', stat['count']) for stat in country_stats],
        )
        _write_distribution_sheet(
            workbook,
            'Tanzania Regions',
            'Region (Tanzania)',
            [(stat['region'] or 'Not Specified', stat['count']) for stat in tanzania_regions],
        )
        _write_distribution_sheet(
            workbook,
            'Dar es Salaam Areas',
            'Center/Area (Dar es Salaam)',
            [(stat['center_area'] or 'Not Specified', stat['count']) for stat in dar_areas],
        )

    if export_type == 'monthly':
        monthly_sheet = workbook.create_sheet(title='Monthly Trend')
        _style_title(monthly_sheet, 'MONTHLY REGISTRATION TREND')
        monthly_sheet['A3'] = 'Month'
        monthly_sheet['B3'] = 'Registrations'
        _style_table_header(monthly_sheet[3])

        current = timezone.now().date().replace(day=1)
        months = []
        for _ in range(12):
            months.insert(0, current)
            if current.month == 1:
                current = current.replace(year=current.year - 1, month=12)
            else:
                current = current.replace(month=current.month - 1)

        row_idx = 4
        for month_start in months:
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
            count = members_qs.filter(
                created_at__date__gte=month_start,
                created_at__date__lt=month_end,
            ).count()
            monthly_sheet.cell(row=row_idx, column=1, value=month_start.strftime('%b %Y'))
            monthly_sheet.cell(row=row_idx, column=2, value=count)
            row_idx += 1

        monthly_sheet.freeze_panes = 'A4'
        _autosize_columns(monthly_sheet)

    summary_sheet.freeze_panes = 'A7'
    _autosize_columns(summary_sheet)

    filename = f'{export_type}-report-{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return _workbook_response(workbook, filename)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_user_activity(request):
    """Export user activity logs"""
    format_type = request.data.get('format', 'csv')
    date_range = request.data.get('date_range', {})
    user_ids = request.data.get('user_ids', [])
    
    # Build queryset
    queryset = AuditLog.objects.all()
    
    if date_range.get('start_date'):
        queryset = queryset.filter(timestamp__gte=date_range['start_date'])
    
    if date_range.get('end_date'):
        queryset = queryset.filter(timestamp__lte=date_range['end_date'])
    
    if user_ids:
        queryset = queryset.filter(user_id__in=user_ids)
    
    if format_type == 'csv':
        response = export_user_activity_csv(queryset)
    elif format_type == 'excel':
        response = export_user_activity_excel(queryset)
    else:
        return Response({'error': 'Unsupported format'}, status=400)
    
    # Log export activity
    file_size = f"{len(response.content) / 1024:.1f} KB"
    ExportHistory.objects.create(
        export_type='users',
        format=format_type,
        created_by=request.user,
        file_size=file_size,
        filters_applied={'date_range': date_range, 'user_ids': user_ids}
    )
    
    return response


def export_user_activity_csv(queryset):
    """Export user activity to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="user_activity_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['User', 'Action', 'Resource Type', 'Resource ID', 'Timestamp', 'IP Address'])
    
    for log in queryset:
        writer.writerow([
            log.user.username if log.user else 'N/A',
            log.action,
            log.resource_type,
            log.resource_id,
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.ip_address
        ])
    
    return response


def export_user_activity_excel(queryset):
    """Export user activity to a styled Excel worksheet."""
    workbook, sheet = _create_workbook('User Activity')
    _style_title(sheet, 'SYSTEM USER ACTIVITY REPORT')

    sheet['A3'] = 'Generated On'
    sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet['A4'] = 'Total Events'
    sheet['B4'] = queryset.count()
    sheet['A3'].font = Font(bold=True)
    sheet['A4'].font = Font(bold=True)

    headers = ['User', 'Action', 'Resource Type', 'Resource ID', 'Timestamp', 'IP Address']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=col, value=header)
    _style_table_header(sheet[6])

    row_idx = 7
    for log in queryset:
        values = [
            log.user.username if log.user else 'N/A',
            log.action,
            log.resource_type,
            log.resource_id,
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.ip_address,
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    sheet.freeze_panes = 'A7'
    _autosize_columns(sheet)

    filename = f'user-activity-{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return _workbook_response(workbook, filename)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_financial(request):
    """Export financial data (placeholder)"""
    format_type = request.data.get('format', 'excel')
    date_range = request.data.get('date_range', {})

    if format_type == 'pdf':
        response = export_financial_pdf(date_range)
    elif format_type == 'excel':
        response = export_financial_excel(date_range)
    elif format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="financial-summary-{datetime.now().strftime("%Y-%m-%d")}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Date', 'Type', 'Amount', 'Member', 'Description'])
        writer.writerow([datetime.now().strftime('%Y-%m-%d'), 'Tithe', '100.00', 'Sample Member', 'Monthly tithe'])
    else:
        return Response({'error': 'Unsupported format'}, status=400)
    
    # Log export activity
    file_size = f"{len(response.content) / 1024:.1f} KB"
    ExportHistory.objects.create(
        export_type='financial',
        format=format_type,
        created_by=request.user,
        file_size=file_size,
        filters_applied={'date_range': date_range}
    )
    
    return response


def export_financial_excel(date_range):
    """Export sample financial summary to executive-style Excel."""
    workbook, sheet = _create_workbook('Financial Summary')
    _style_title(sheet, 'KUSANYIKO FINANCIAL SUMMARY REPORT')

    start_label = date_range.get('start_date') or 'Beginning'
    end_label = date_range.get('end_date') or 'Present'
    sheet['A3'] = 'Prepared On'
    sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet['A4'] = 'Reporting Window'
    sheet['B4'] = f'{start_label} to {end_label}'
    sheet['A3'].font = Font(bold=True)
    sheet['A4'].font = Font(bold=True)

    headers = ['Date', 'Type', 'Amount (TZS)', 'Member', 'Description']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=col, value=header)
    _style_table_header(sheet[6])

    sample_rows = [
        [datetime.now().strftime('%Y-%m-%d'), 'Tithe', 100000, 'Sample Member', 'Monthly tithe'],
        [datetime.now().strftime('%Y-%m-%d'), 'Offering', 50000, 'Sample Member', 'Sunday service offering'],
        [datetime.now().strftime('%Y-%m-%d'), 'Special Seed', 250000, 'Sample Member', 'Special contribution'],
    ]

    row_idx = 7
    for row in sample_rows:
        for col, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col, value=value)
        sheet.cell(row=row_idx, column=3).number_format = '#,##0'
        row_idx += 1

    sheet.freeze_panes = 'A7'
    _autosize_columns(sheet)

    filename = f'financial-summary-{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return _workbook_response(workbook, filename)


def export_financial_pdf(date_range):
    """Export sample financial summary to branded executive PDF."""
    styles = _pdf_styles()
    start_label = date_range.get('start_date') or 'Beginning'
    end_label = date_range.get('end_date') or 'Present'

    sample_rows = [
        [datetime.now().strftime('%Y-%m-%d'), 'Tithe', 100000, 'Sample Member', 'Monthly tithe'],
        [datetime.now().strftime('%Y-%m-%d'), 'Offering', 50000, 'Sample Member', 'Sunday service offering'],
        [datetime.now().strftime('%Y-%m-%d'), 'Special Seed', 250000, 'Sample Member', 'Special contribution'],
    ]
    total_amount = sum(row[2] for row in sample_rows)

    story = []
    story.extend(
        _pdf_header_block(
            'KUSANYIKO FINANCIAL SUMMARY REPORT',
            f'Reporting window: {start_label} to {end_label}',
        )
    )
    story.append(Paragraph('Executive KPI Summary', styles['ExecutiveSection']))
    story.append(
        _pdf_table(
            [
                ['Indicator', 'Value'],
                ['Total Transactions (Sample)', len(sample_rows)],
                ['Total Amount (TZS)', f'{total_amount:,}'],
            ],
            col_widths=[10 * cm, 5.5 * cm],
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph('Transaction Register', styles['ExecutiveSection']))
    transaction_rows = [['Date', 'Type', 'Amount (TZS)', 'Member', 'Description']]
    for row in sample_rows:
        transaction_rows.append([row[0], row[1], f'{row[2]:,}', row[3], row[4]])

    story.append(
        _pdf_table(
            transaction_rows,
            col_widths=[2.4 * cm, 2.4 * cm, 2.8 * cm, 3.5 * cm, 5.4 * cm],
        )
    )

    filename = f'financial-summary-{datetime.now().strftime("%Y-%m-%d")}.pdf'
    return _pdf_response(story, filename)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_history(request):
    """Get export history"""
    exports = ExportHistory.objects.filter(created_by=request.user).order_by('-created_at')[:20]
    
    export_data = []
    for export in exports:
        export_data.append({
            'id': export.id,
            'export_type': export.export_type,
            'format': export.format,
            'created_at': export.created_at,
            'file_size': export.file_size,
            'download_count': export.download_count,
            'created_by': {
                'id': export.created_by.id,
                'username': export.created_by.username
            }
        })
    
    return Response(export_data)
